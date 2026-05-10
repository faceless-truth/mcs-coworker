"""
Chattel mortgage amortisation schedule generator.

Implements the AASB 9 effective-interest-method approach described in
specialists/prompts/chattel_mortgage.md. Convention is monthly payments
in advance unless explicitly stated otherwise.
"""

from __future__ import annotations
from dataclasses import dataclass
from datetime import date
from io import BytesIO

from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


@dataclass
class LoanInputs:
    asset: str
    start_date: date
    amount_financed: float
    stated_rate_pa: float          # nominal annual % as disclosed
    term_months: int
    monthly_payment: float
    balloon: float = 0.0
    convention: str = "in_advance"  # "in_advance" | "in_arrears"


@dataclass
class ScheduleRow:
    period: int
    date: date
    payment: float
    interest: float
    principal: float
    balance: float


# ---------- core math ----------

def _run_schedule(
    principal: float,
    payment: float,
    n: int,
    monthly_rate: float,
    balloon: float,
    start_date: date,
    convention: str,
) -> tuple[list[ScheduleRow], float]:
    """
    Run the schedule forward at a given monthly rate.
    Returns (rows, final_balance). Used both by the solver and the final build.
    """
    rows: list[ScheduleRow] = []
    balance = principal

    if convention == "in_advance":
        # Payment 1 immediately at start_date — zero interest, full principal
        balance -= payment
        rows.append(ScheduleRow(1, start_date, payment, 0.0, payment, balance))
        for i in range(2, n + 1):
            interest = balance * monthly_rate
            principal_portion = payment - interest
            balance -= principal_portion
            rows.append(ScheduleRow(
                i, start_date + relativedelta(months=i - 1),
                payment, interest, principal_portion, balance,
            ))
    elif convention == "in_arrears":
        for i in range(1, n + 1):
            interest = balance * monthly_rate
            principal_portion = payment - interest
            balance -= principal_portion
            rows.append(ScheduleRow(
                i, start_date + relativedelta(months=i - 1),
                payment, interest, principal_portion, balance,
            ))
    else:
        raise ValueError(f"Unknown convention: {convention}")

    if balloon > 0:
        interest = balance * monthly_rate
        principal_portion = balloon - interest
        balance -= principal_portion
        rows.append(ScheduleRow(
            n + 1, start_date + relativedelta(months=n),
            balloon, interest, principal_portion, balance,
        ))

    return rows, balance


def solve_effective_monthly_rate(
    principal: float,
    payment: float,
    n: int,
    balloon: float,
    start_date: date,
    convention: str,
    tol: float = 1e-9,
    max_iter: int = 200,
) -> float:
    """
    Bisect for the monthly rate that makes the final balance zero.
    Higher rate → higher final balance, so:
       final > 0  → rate too high → tighten upper bound
       final < 0  → rate too low  → tighten lower bound
    """
    lo, hi = 0.0, 0.5  # 0% to 50% per month brackets every plausible chattel rate
    mid = (lo + hi) / 2.0
    for _ in range(max_iter):
        mid = (lo + hi) / 2.0
        _, final = _run_schedule(principal, payment, n, mid, balloon, start_date, convention)
        if abs(final) < tol:
            return mid
        if final > 0:
            hi = mid
        else:
            lo = mid
    return mid


def build_loan_schedule(loan: LoanInputs) -> tuple[list[ScheduleRow], float]:
    """
    Solve for the effective rate and return the final schedule + monthly effective rate.
    """
    monthly_rate = solve_effective_monthly_rate(
        principal=loan.amount_financed,
        payment=loan.monthly_payment,
        n=loan.term_months,
        balloon=loan.balloon,
        start_date=loan.start_date,
        convention=loan.convention,
    )
    rows, _final = _run_schedule(
        loan.amount_financed,
        loan.monthly_payment,
        loan.term_months,
        monthly_rate,
        loan.balloon,
        loan.start_date,
        loan.convention,
    )
    return rows, monthly_rate


# ---------- FY grouping ----------

def au_fy(d: date) -> int:
    """Return the AU financial year number (e.g. FY24 = 1 Jul 2023 – 30 Jun 2024 → 24)."""
    fy_full = d.year + 1 if d.month >= 7 else d.year
    return fy_full % 100


# ---------- xlsx render ----------

CURRENCY_FORMAT = '$#,##0.00;($#,##0.00);"-"'

HEADER_FILL = PatternFill("solid", fgColor="305496")
FY_SUBTOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
GRAND_TOTAL_FILL = PatternFill("solid", fgColor="8EA9DB")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", bold=True, size=13)
SUBTITLE_FONT = Font(name="Calibri", italic=True, size=10)
GRAND_TOTAL_FONT = Font(name="Calibri", bold=True)

THIN = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MEDIUM = Side(style="medium", color="000000")
GT_BORDER = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)


def _safe_sheet_title(name: str, used: set[str]) -> str:
    """Excel sheet titles are <=31 chars and can't contain :\\/?*[]"""
    cleaned = "".join(c for c in name if c not in r':\/?*[]')[:31].strip() or "Schedule"
    base = cleaned
    n = 2
    while cleaned in used:
        suffix = f" ({n})"
        cleaned = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(cleaned)
    return cleaned


def render_workbook(loans: list[LoanInputs]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    used_titles: set[str] = set()

    for loan in loans:
        rows, monthly_rate = build_loan_schedule(loan)
        effective_rate_pa = monthly_rate * 12 * 100
        ws = wb.create_sheet(_safe_sheet_title(loan.asset, used_titles))

        # Title block
        ws["A1"] = loan.asset
        ws["A1"].font = TITLE_FONT
        ws.merge_cells("A1:F1")
        balloon_str = f"${loan.balloon:,.2f}" if loan.balloon else "no balloon"
        ws["A2"] = (
            f"Amount financed ${loan.amount_financed:,.2f}  •  "
            f"Stated {loan.stated_rate_pa:.2f}% p.a.  •  "
            f"Effective {effective_rate_pa:.3f}% p.a.  •  "
            f"{loan.term_months} months  •  "
            f"Monthly payment ${loan.monthly_payment:,.2f}  •  "
            f"First payment {loan.start_date.strftime('%d/%m/%Y')}  •  "
            f"{balloon_str}"
        )
        ws["A2"].font = SUBTITLE_FONT
        ws.merge_cells("A2:F2")

        # Column header (row 4)
        headers = ["#", "Date", "Payment", "Interest", "Principal", "Balance"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        ws.freeze_panes = "A5"

        # Group rows by FY, write data + FY subtotals + grand total
        current_fy = None
        fy_block_start_row: int | None = None
        fy_subtotal_rows: list[int] = []
        excel_row = 5

        def flush_fy_subtotal(end_row: int):
            nonlocal excel_row
            sub_row = excel_row
            ws.cell(row=sub_row, column=2,
                    value=f"FY{current_fy:02d} Total").alignment = Alignment(horizontal="right")
            for col_letter, col_idx in [("C", 3), ("D", 4), ("E", 5)]:
                cell = ws.cell(row=sub_row, column=col_idx,
                               value=f"=SUM({col_letter}{fy_block_start_row}:{col_letter}{end_row})")
                cell.number_format = CURRENCY_FORMAT
            for col_idx in range(1, 7):
                c = ws.cell(row=sub_row, column=col_idx)
                c.fill = FY_SUBTOTAL_FILL
                c.font = Font(name="Calibri", bold=True)
                c.border = THIN_BORDER
            fy_subtotal_rows.append(sub_row)
            excel_row += 1
            # Spacer
            excel_row += 1

        for r in rows:
            row_fy = au_fy(r.date)
            if current_fy is None:
                current_fy = row_fy
                fy_block_start_row = excel_row
            elif row_fy != current_fy:
                flush_fy_subtotal(excel_row - 1)
                current_fy = row_fy
                fy_block_start_row = excel_row

            ws.cell(row=excel_row, column=1, value=r.period).alignment = Alignment(horizontal="center")
            date_cell = ws.cell(row=excel_row, column=2, value=r.date)
            date_cell.number_format = "dd/mm/yyyy"
            date_cell.alignment = Alignment(horizontal="center")
            for col_idx, val in [(3, r.payment), (4, r.interest), (5, r.principal), (6, r.balance)]:
                c = ws.cell(row=excel_row, column=col_idx, value=val)
                c.number_format = CURRENCY_FORMAT
            for col_idx in range(1, 7):
                ws.cell(row=excel_row, column=col_idx).border = THIN_BORDER
            excel_row += 1

        # Final FY subtotal
        if current_fy is not None:
            flush_fy_subtotal(excel_row - 1)

        # Grand total
        gt_row = excel_row
        ws.cell(row=gt_row, column=2,
                value="GRAND TOTAL").alignment = Alignment(horizontal="right")
        for col_letter, col_idx in [("C", 3), ("D", 4), ("E", 5)]:
            refs = ",".join(f"{col_letter}{r}" for r in fy_subtotal_rows)
            cell = ws.cell(row=gt_row, column=col_idx, value=f"=SUM({refs})")
            cell.number_format = CURRENCY_FORMAT
        for col_idx in range(1, 7):
            c = ws.cell(row=gt_row, column=col_idx)
            c.fill = GRAND_TOTAL_FILL
            c.font = GRAND_TOTAL_FONT
            c.border = GT_BORDER

        # Column widths
        widths = {"A": 6, "B": 14, "C": 14, "D": 14, "E": 14, "F": 16}
        for letter, w in widths.items():
            ws.column_dimensions[letter].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
