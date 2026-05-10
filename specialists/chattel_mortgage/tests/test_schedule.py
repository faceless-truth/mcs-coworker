"""
Tests for the chattel mortgage solver, in-advance/in-arrears amortisation,
balloon handling, AU FY grouping, and the .xlsx renderer.
"""
from datetime import date
from io import BytesIO

import sys
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(REPO_ROOT))

from openpyxl import load_workbook  # noqa: E402

from specialists.chattel_mortgage.schedule import (  # noqa: E402
    LoanInputs,
    _run_schedule,
    au_fy,
    build_loan_schedule,
    render_workbook,
    solve_effective_monthly_rate,
)


def test_au_fy():
    assert au_fy(date(2023, 10, 5)) == 24
    assert au_fy(date(2026, 6, 10)) == 26
    assert au_fy(date(2026, 7, 1)) == 27


def test_in_advance_payment_1_zero_interest():
    loan = LoanInputs(
        asset="Test", start_date=date(2024, 1, 1),
        amount_financed=50000, stated_rate_pa=6.5, term_months=60,
        monthly_payment=978.31, balloon=0,
    )
    rows, _ = build_loan_schedule(loan)
    assert rows[0].interest == 0.0
    assert abs(rows[0].principal - 978.31) < 1e-6


def test_schedule_amortises_to_zero_no_balloon():
    loan = LoanInputs(
        asset="A", start_date=date(2024, 1, 1),
        amount_financed=50000, stated_rate_pa=6.5, term_months=60,
        monthly_payment=975.00, balloon=0,
    )
    rows, _ = build_loan_schedule(loan)
    assert abs(rows[-1].balance) < 0.01


def test_schedule_amortises_to_zero_with_balloon():
    loan = LoanInputs(
        asset="A", start_date=date(2024, 1, 1),
        amount_financed=60000, stated_rate_pa=6.5, term_months=60,
        monthly_payment=850.00, balloon=10000,
    )
    rows, _ = build_loan_schedule(loan)
    assert rows[-1].period == 61               # balloon row exists
    assert abs(rows[-1].balance) < 0.01


def test_solver_recovers_known_rate():
    # Construct a payment from a known rate, verify solver returns it.
    principal, n, balloon = 50000, 60, 0
    target_monthly = 0.005   # 6% p.a. nominal
    # Solve PMT for in-advance with no balloon by bisecting on payment instead.
    lo, hi = 100.0, 5000.0
    for _ in range(200):
        mid = (lo + hi) / 2
        _, final = _run_schedule(principal, mid, n, target_monthly, balloon,
                                 date(2024, 1, 1), "in_advance")
        if final > 0:
            lo = mid
        else:
            hi = mid
    payment = (lo + hi) / 2
    recovered = solve_effective_monthly_rate(
        principal, payment, n, balloon, date(2024, 1, 1), "in_advance",
    )
    assert abs(recovered - target_monthly) < 1e-6


def test_workbook_has_fy_subtotals_and_grand_total():
    loan = LoanInputs(
        asset="2024 Test Truck", start_date=date(2024, 3, 1),
        amount_financed=80000, stated_rate_pa=7.0, term_months=48,
        monthly_payment=1950.00, balloon=0,
    )
    blob = render_workbook([loan])
    wb = load_workbook(BytesIO(blob))
    ws = wb["2024 Test Truck"]
    cells = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
    assert any(
        isinstance(v, str) and v.startswith("FY") and v.endswith("Total")
        for v in cells
    )
    assert any(v == "GRAND TOTAL" for v in cells)


def test_multi_loan_workbook_has_one_sheet_per_loan():
    loans = [
        LoanInputs("Asset One", date(2024, 1, 1), 50000, 6.5, 60, 978.0, 0),
        LoanInputs("Asset Two", date(2025, 7, 1), 30000, 7.0, 36, 925.0, 0),
    ]
    blob = render_workbook(loans)
    wb = load_workbook(BytesIO(blob))
    assert "Asset One" in wb.sheetnames
    assert "Asset Two" in wb.sheetnames


def test_in_arrears_payment_1_charges_interest():
    """In arrears, period 1 accrues a month of interest before the first
    payment, unlike in_advance where the first payment is pure principal."""
    loan = LoanInputs(
        asset="A", start_date=date(2024, 1, 1),
        amount_financed=50000, stated_rate_pa=6.5, term_months=60,
        monthly_payment=978.00, balloon=0,
        convention="in_arrears",
    )
    rows, _ = build_loan_schedule(loan)
    assert rows[0].interest > 0


def test_subtotal_uses_excel_sum_formula_not_a_constant():
    """Spec: FY subtotals must be Excel =SUM(...) formulas so the user can
    see the cell-range provenance, not a baked-in number."""
    loan = LoanInputs(
        asset="Formula Test", start_date=date(2024, 3, 1),
        amount_financed=60000, stated_rate_pa=7.0, term_months=36,
        monthly_payment=1850.00, balloon=0,
    )
    blob = render_workbook([loan])
    wb = load_workbook(BytesIO(blob))
    ws = wb["Formula Test"]
    # Find the first FY subtotal row by scanning column B for "FY..Total"
    found_formula = False
    for r in range(1, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value
        if isinstance(b, str) and b.startswith("FY") and b.endswith("Total"):
            for col_idx in (3, 4, 5):
                v = ws.cell(row=r, column=col_idx).value
                assert isinstance(v, str) and v.startswith("=SUM("), (
                    f"FY subtotal col {col_idx} should be a SUM formula, got {v!r}"
                )
                found_formula = True
            break
    assert found_formula, "no FY subtotal row found in workbook"
